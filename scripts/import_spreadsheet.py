#!/usr/bin/env python3
"""Import historical WC26 data from the scoring spreadsheet into scoreboard.json.

Strategy (confirmed with user): mirror the spreadsheet's already-computed
per-stage point totals as a FROZEN historical baseline (group stage -> R16).
QF -> Final are computed live in the app from matches + predictions.
"""
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "WC 2026 SCORING TABLE.xlsx"

# LEADERBOARD columns (1-indexed)
COL = {
    "pos": 1, "prv": 3, "name": 5, "final": 6,
    "dgs_group": 7, "dgs_knockout": 8,
    "goalscorerPoints": 9, "overallScore": 10,
    "matchday1": 11, "matchday2": 12, "matchday3": 13,
    "groupTotal": 14, "r32": 15, "r16": 16, "qf": 17, "sf": 18, "final_pts": 19,
    "overallPoints": 20,
}

# Only these real participants (rest of the sheet rows are template placeholders).
def is_real_name(name):
    if not name:
        return False
    name = str(name).strip()
    # Placeholder rows are single/double letters (L, M, ... AA, BB) or 'x'.
    if len(name) <= 2:
        return False
    if name.lower() == "x":
        return False
    return " " in name  # real names are "FIRST LAST"


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else 0


def split_finalists(text):
    if not text or str(text).strip() in ("-", ""):
        return []
    parts = re.split(r"\s+[vV]\s+", str(text).strip())
    return [p.strip() for p in parts if p.strip()]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["LEADERBOARD"]

    players = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=COL["name"]).value
        if not is_real_name(name):
            continue
        name = str(name).strip()

        final_text = ws.cell(row=r, column=COL["final"]).value
        baseline = {
            "matchday1": int(num(ws.cell(row=r, column=COL["matchday1"]).value)),
            "matchday2": int(num(ws.cell(row=r, column=COL["matchday2"]).value)),
            "matchday3": int(num(ws.cell(row=r, column=COL["matchday3"]).value)),
            "groupTotal": int(num(ws.cell(row=r, column=COL["groupTotal"]).value)),
            "r32": int(num(ws.cell(row=r, column=COL["r32"]).value)),
            "r16": int(num(ws.cell(row=r, column=COL["r16"]).value)),
            "goalscorerPoints": int(num(ws.cell(row=r, column=COL["goalscorerPoints"]).value)),
        }
        prv = ws.cell(row=r, column=COL["prv"]).value
        players.append({
            "name": name,
            "previousPosition": int(prv) if isinstance(prv, (int, float)) and not isinstance(prv, bool) else None,
            "finalPrediction": str(final_text).strip() if final_text else "",
            "predictedFinalists": split_finalists(final_text),
            "designatedGoalscorerGroup": str(ws.cell(row=r, column=COL["dgs_group"]).value or "").strip(),
            "designatedGoalscorerKnockout": str(ws.cell(row=r, column=COL["dgs_knockout"]).value or "").strip(),
            "baseline": baseline,
            "predictions": {},
        })

    data = {
        "lastUpdated": datetime.now(timezone.utc).isoformat(),
        "actualFinalists": [],
        "matches": [],  # QF -> Final added here for live scoring
        "players": players,
    }

    # Sanity check: overall from baseline should equal sheet's OVERALL SCORE (col J).
    print(f"Imported {len(players)} players:", file=sys.stderr)
    for r, p in zip(range(3, 3 + len(players)), players):
        b = p["baseline"]
        overall = b["groupTotal"] + b["r32"] + b["r16"] + b["goalscorerPoints"]
        sheet_overall = num(ws.cell(row=r, column=COL["overallScore"]).value)
        flag = "OK" if overall == sheet_overall else f"MISMATCH sheet={sheet_overall}"
        print(f"  {p['name']:22} overall={overall:4} [{flag}]", file=sys.stderr)

    out = ROOT / "server" / "seed" / "scoreboard.json"
    out.write_text(json.dumps(data, indent=2) + "\n")
    print(f"\nWrote {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
